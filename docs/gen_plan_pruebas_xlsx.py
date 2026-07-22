# -*- coding: utf-8 -*-
"""
Genera plan_pruebas_manual.xlsx a partir de doc/plan_pruebas_manual.md.

El MD es la ÚNICA fuente de verdad. Estructura jerárquica:
    Módulo (## N. …) → Caso (### CU-NN · título) → Paso (fila CU-NN-n).

Produce un libro con dos pestañas:
  · "Casos de prueba" — un PASO por fila, con columnas de ejecución:
        Resultado obtenido | Estado (Pendiente/OK/KO) | Bugs (id. ticketing) | Evidencia
  · "Resumen" — KPIs y desglose por módulo y por caso con fórmulas vivas
        (COUNTIF/COUNTIFS) + gráficos. Recalcula solo al rellenar Estado.

Uso:  python doc/gen_plan_pruebas_xlsx.py
"""
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# Colores de estado para las gráficas (OK verde, KO rojo, Pendiente gris)
CLR_OK, CLR_KO, CLR_PEND = "00B050", "FF0000", "BFBFBF"

BASE = Path(__file__).resolve().parent
MD = BASE / "plan_pruebas_manual.md"
OUT = BASE / "plan_pruebas_manual.xlsx"

CASO_RE = re.compile(r"^###\s+(CU-\d+)\s+·\s+(.*)$")
STEP_RE = re.compile(r"^[A-Z]+-\d+-\d+$")          # CU-01-3
PIPE = re.compile(r"(?<!\\)\|")


def clean(text: str) -> str:
    text = text.replace("\\_", "_").replace("\\*", "*").replace("\\|", "|")
    return text.replace("**", "").replace("`", "").strip()


def cells_of(line: str):
    body = line.strip()
    if body.startswith("|"):
        body = body[1:]
    if body.endswith("|"):
        body = body[:-1]
    return [c.strip() for c in PIPE.split(body)]


def parse(md_text: str):
    """Lista de pasos: {modulo, caso_id, caso, paso, area, desc, cond, esperado}."""
    pasos = []
    modulo = "(sin módulo)"
    caso_id = caso = None
    for raw in md_text.splitlines():
        s = raw.strip()
        if s.startswith("## ") and not s.startswith("###"):
            modulo = re.sub(r"^\d+\.\s*", "", s[3:].strip())
            caso_id = caso = None
            continue
        m = CASO_RE.match(s)
        if m:
            caso_id, caso = m.group(1), clean(m.group(2))
            continue
        if s.startswith("|") and caso_id:
            c = cells_of(s)
            if not c or not STEP_RE.match(clean(c[0])):
                continue
            v = [clean(x) for x in c]
            while len(v) < 5:
                v.append("")
            if len(v) > 5:
                v = v[:4] + [" | ".join(v[4:])]
            pasos.append({
                "modulo": modulo, "caso_id": caso_id, "caso": caso,
                "paso": v[0], "area": v[1], "desc": v[2],
                "cond": v[3], "esperado": v[4],
            })
    return pasos


# Columnas de ejecución que se preservan al regenerar (modo merge).
KEEP_COLS = ["Resultado obtenido", "Estado", "Bugs", "Evidencia", "Asignado a"]


def read_existing(path):
    """Lee el .xlsx previo y devuelve {paso_id: {col: valor}} de las columnas
    de ejecución ya rellenadas, para no perderlas al regenerar. Casa por `Paso`
    leyendo las columnas por su CABECERA (robusto si cambió el nº de columnas)."""
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return {}
    if "Casos de prueba" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Casos de prueba"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return {}
    idx = {name: i for i, name in enumerate(header) if name}
    if "Paso" not in idx:
        wb.close()
        return {}
    pcol = idx["Paso"]
    saved = {}
    for r in rows:
        paso = r[pcol] if pcol < len(r) else None
        if not paso:
            continue
        saved[str(paso).strip()] = {
            c: (r[idx[c]] if c in idx and idx[c] < len(r) else None)
            for c in KEEP_COLS
        }
    wb.close()
    return saved


# ── Estilos ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(bold=True, size=12, color="1F4E78")
KPI_LABEL = Font(bold=True, size=11)
KPI_FILL = PatternFill("solid", fgColor="DDEBF7")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

GREEN, GREEN_F = PatternFill("solid", fgColor="C6EFCE"), Font(color="006100")
RED, RED_F = PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006")
GREY, GREY_F = PatternFill("solid", fgColor="EDEDED"), Font(color="808080")

COLS = ["Módulo", "ID-CASO", "Caso", "Paso", "Área", "Descripción",
        "Condiciones / Datos", "Resultado esperado", "Resultado obtenido",
        "Estado", "Bugs", "Evidencia", "Asignado a"]
WIDTHS = [24, 9, 34, 11, 20, 50, 42, 46, 36, 12, 14, 22, 18]
STATE_COL = "J"   # Estado
DATA = "'Casos de prueba'"


def build_cases_sheet(ws, pasos, saved=None):
    saved = saved or {}
    ws.title = "Casos de prueba"
    for i, name in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]

    for r, p in enumerate(pasos, start=2):
        prev = saved.get(p["paso"], {})
        row = [p["modulo"], p["caso_id"], p["caso"], p["paso"], p["area"],
               p["desc"], p["cond"], p["esperado"],
               prev.get("Resultado obtenido") or "",
               prev.get("Estado") or "Pendiente",
               prev.get("Bugs") or "",
               prev.get("Evidencia") or "",
               prev.get("Asignado a") or ""]
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.border = BORDER
            c.alignment = CENTER if ci in (2, 4, 10) else WRAP_TOP
            if ci == 10:
                c.font = Font(bold=True)

    last = len(pasos) + 1
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:M{last}"

    dv = DataValidation(type="list", formula1='"Pendiente,OK,KO"', allow_blank=False)
    dv.error, dv.errorTitle = "Valores: Pendiente, OK, KO", "Estado no válido"
    ws.add_data_validation(dv)
    dv.add(f"{STATE_COL}2:{STATE_COL}{last}")

    rng = f"{STATE_COL}2:{STATE_COL}{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OK"'], fill=GREEN, font=GREEN_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"KO"'], fill=RED, font=RED_F))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pendiente"'], fill=GREY, font=GREY_F))
    return last


def _hdr(ws, row, headers, start=1):
    for ci, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill, c.font, c.border, c.alignment = HEADER_FILL, HEADER_FONT, BORDER, CENTER


def build_summary_sheet(ws, pasos, last):
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["B"].width = 36  # Caso (en tabla por caso)

    ws["A1"] = "Plan de Pruebas — Ofertas Hipotecarias · Seguimiento de ejecución"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    S = f"{DATA}!{STATE_COL}2:{STATE_COL}{last}"
    B = f"{DATA}!$B$2:$B${last}"          # ID-CASO
    A = f"{DATA}!$A$2:$A${last}"          # Módulo
    SS = f"{DATA}!${STATE_COL}$2:${STATE_COL}${last}"

    ws["A3"] = "Indicadores globales (pasos)"
    ws["A3"].font = SUB_FONT
    kpis = [
        ("Total de pasos",          f"=COUNTA({DATA}!D2:D{last})"),
        ("Pendientes",              f'=COUNTIF({S},"Pendiente")'),
        ("OK",                      f'=COUNTIF({S},"OK")'),
        ("KO (incidencias)",        f'=COUNTIF({S},"KO")'),
        ("% Avance (ejecutados)",   f'=IFERROR((COUNTIF({S},"OK")+COUNTIF({S},"KO"))/COUNTA({DATA}!D2:D{last}),0)'),
        ("% Éxito (OK/ejecutados)", f'=IFERROR(COUNTIF({S},"OK")/(COUNTIF({S},"OK")+COUNTIF({S},"KO")),0)'),
        ("Bugs referenciados",      f'=SUMPRODUCT(--({DATA}!K2:K{last}<>""))'),
    ]
    r = 4
    for label, formula in kpis:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font, lc.fill, lc.border = KPI_LABEL, KPI_FILL, BORDER
        vc = ws.cell(row=r, column=2, value=formula)
        vc.border, vc.alignment = BORDER, CENTER
        if label.startswith("%"):
            vc.number_format = "0.0%"
        r += 1

    # Tabla auxiliar para gráfico de estado
    ws["D3"], ws["E3"] = "Estado", "Pasos"
    ws["D3"].font = ws["E3"].font = Font(bold=True)
    for i, (e, f) in enumerate([("Pendiente", f'=COUNTIF({S},"Pendiente")'),
                                ("OK", f'=COUNTIF({S},"OK")'),
                                ("KO", f'=COUNTIF({S},"KO")')]):
        ws.cell(row=4 + i, column=4, value=e).border = BORDER
        vc = ws.cell(row=4 + i, column=5, value=f)
        vc.border, vc.alignment = BORDER, CENTER

    pie = PieChart()
    pie.title = "Distribución por estado"
    pie.add_data(Reference(ws, min_col=5, min_row=3, max_row=6), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=4, min_row=4, max_row=6))
    pie.height, pie.width = 7, 11
    # Color por porción: Pendiente=gris, OK=verde, KO=rojo (orden de filas 4,5,6)
    pie.series[0].data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(solidFill=clr))
        for i, clr in enumerate((CLR_PEND, CLR_OK, CLR_KO))
    ]
    ws.add_chart(pie, "G3")

    # ── Desglose por módulo ──
    modulos = []
    for p in pasos:
        if p["modulo"] not in modulos:
            modulos.append(p["modulo"])

    mrow = r + 1
    ws.cell(row=mrow, column=1, value="Desglose por módulo").font = SUB_FONT
    mh = mrow + 1
    _hdr(ws, mh, ["Módulo", "Total", "Pendiente", "OK", "KO", "% Avance"])
    row = mh + 1
    for mod in modulos:
        ws.cell(row=row, column=1, value=mod).border = BORDER
        ws.cell(row=row, column=1).alignment = WRAP_TOP
        ws.cell(row=row, column=2, value=f'=COUNTIF({A},A{row})').border = BORDER
        ws.cell(row=row, column=3, value=f'=COUNTIFS({A},A{row},{SS},"Pendiente")').border = BORDER
        ws.cell(row=row, column=4, value=f'=COUNTIFS({A},A{row},{SS},"OK")').border = BORDER
        ws.cell(row=row, column=5, value=f'=COUNTIFS({A},A{row},{SS},"KO")').border = BORDER
        pc = ws.cell(row=row, column=6, value=f'=IFERROR((D{row}+E{row})/B{row},0)')
        pc.number_format, pc.border = "0.0%", BORDER
        for ci in range(2, 7):
            ws.cell(row=row, column=ci).alignment = CENTER
        if (row - mh) % 2 == 0:
            for ci in range(1, 7):
                ws.cell(row=row, column=ci).fill = ALT_FILL
        row += 1
    mend = row - 1

    bar = BarChart()
    bar.type, bar.grouping, bar.overlap = "bar", "stacked", 100
    bar.title = "Avance y errores por módulo"
    bar.height, bar.width = max(8, 0.5 * len(modulos) + 3), 18
    bar.add_data(Reference(ws, min_col=3, max_col=5, min_row=mh, max_row=mend), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=mh + 1, max_row=mend))
    # Series Pendiente=gris, OK=verde, KO=rojo
    for serie, clr in zip(bar.series, (CLR_PEND, CLR_OK, CLR_KO)):
        serie.graphicalProperties = GraphicalProperties(solidFill=clr)
    ws.add_chart(bar, f"H{mrow}")

    # ── Desglose por caso ──
    casos = []
    seen = set()
    for p in pasos:
        if p["caso_id"] not in seen:
            seen.add(p["caso_id"])
            casos.append((p["caso_id"], p["caso"]))

    crow = mend + 3
    ws.cell(row=crow, column=1, value="Desglose por caso").font = SUB_FONT
    ch = crow + 1
    _hdr(ws, ch, ["ID-CASO", "Caso", "Total", "Pend.", "OK", "KO", "% Avance", "Estado caso"])
    row = ch + 1
    for cid, ctitle in casos:
        ws.cell(row=row, column=1, value=cid).border = BORDER
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=2, value=ctitle).border = BORDER
        ws.cell(row=row, column=2).alignment = WRAP_TOP
        ws.cell(row=row, column=3, value=f'=COUNTIF({B},A{row})').border = BORDER
        ws.cell(row=row, column=4, value=f'=COUNTIFS({B},A{row},{SS},"Pendiente")').border = BORDER
        ws.cell(row=row, column=5, value=f'=COUNTIFS({B},A{row},{SS},"OK")').border = BORDER
        ws.cell(row=row, column=6, value=f'=COUNTIFS({B},A{row},{SS},"KO")').border = BORDER
        pc = ws.cell(row=row, column=7, value=f'=IFERROR((E{row}+F{row})/C{row},0)')
        pc.number_format, pc.border = "0.0%", BORDER
        ec = ws.cell(row=row, column=8, value=f'=IF(F{row}>0,"KO",IF(D{row}>0,"En curso","OK"))')
        ec.border = BORDER
        for ci in (3, 4, 5, 6, 7, 8):
            ws.cell(row=row, column=ci).alignment = CENTER
        if (row - ch) % 2 == 0:
            for ci in range(1, 9):
                ws.cell(row=row, column=ci).fill = ALT_FILL
        row += 1
    cend = row - 1

    # Semáforo del estado-caso
    crng = f"H{ch + 1}:H{cend}"
    ws.conditional_formatting.add(crng, CellIsRule(operator="equal", formula=['"OK"'], fill=GREEN, font=GREEN_F))
    ws.conditional_formatting.add(crng, CellIsRule(operator="equal", formula=['"KO"'], fill=RED, font=RED_F))
    ws.conditional_formatting.add(crng, CellIsRule(operator="equal", formula=['"En curso"'], fill=GREY, font=GREY_F))


def main():
    if not MD.exists():
        sys.exit(f"No se encuentra {MD}")
    pasos = parse(MD.read_text(encoding="utf-8"))
    if not pasos:
        sys.exit("No se extrajo ningún paso del MD.")

    saved = read_existing(OUT)              # merge: conserva ejecución previa por Paso
    ids = {p["paso"] for p in pasos}
    preservados = sum(1 for k in saved if k in ids)
    huerfanos = [k for k in saved if k not in ids]

    wb = Workbook()
    last = build_cases_sheet(wb.active, pasos, saved)
    build_summary_sheet(wb.create_sheet("Resumen"), pasos, last)
    wb.move_sheet("Resumen", -(len(wb.sheetnames) - 1))
    wb.active = 0
    wb.save(OUT)

    from collections import Counter
    ncasos = len({p["caso_id"] for p in pasos})
    print(f"OK - {ncasos} casos / {len(pasos)} pasos -> {OUT.name}")
    if saved:
        print(f"   merge: {preservados} pasos con ejecución preservada", end="")
        if huerfanos:
            print(f"; {len(huerfanos)} pasos del .xlsx previo ya no existen en el MD: {', '.join(sorted(huerfanos)[:8])}{'…' if len(huerfanos) > 8 else ''}")
        else:
            print()
    by_mod = Counter(p["modulo"] for p in pasos)
    for mod, n in by_mod.items():
        print(f"   {n:3d}  {mod}")


if __name__ == "__main__":
    main()
